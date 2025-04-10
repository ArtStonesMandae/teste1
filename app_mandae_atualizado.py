
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import io
import re

st.set_page_config(page_title="Gerador Mandae", layout="centered")

st.title("📦 Gerador de Planilhas Mandaê")
st.write("Faça upload do seu arquivo .csv e baixe a planilha formatada para envio via Mandaê.")

uploaded_file = st.file_uploader("Selecione o arquivo CSV:", type=["csv"])
if uploaded_file:
    try:
        df = pd.read_csv(uploaded_file, encoding='latin1', sep=';', dtype=str)
    except Exception as e:
        st.error(f"Erro ao ler o arquivo: {e}")
        st.stop()

    if df['Destinatário'].isna().any():
        st.error("Existem linhas com DESTINATÁRIO vazio. Corrija antes de continuar.")
        st.stop()

    def format_document(cpf, cnpj):
        if pd.notna(cnpj):
            return re.sub(r'\D', '', cnpj).zfill(14)
        elif pd.notna(cpf):
            return re.sub(r'\D', '', cpf).zfill(11)
        return ''

    def get_phone(row):
        return row['Telefone'] if pd.notna(row['Telefone']) else row['Celular']

    def clean_cep(cep):
        return re.sub(r'\D', '', cep)

    saida_df = pd.DataFrame({
        'NOME DO DESTINATÁRIO*': df['Destinatário'],
        'NOME DA EMPRESA (EM CASO DE ENDEREÇO COMERCIAL)': df['Razão Social'],
        'E-MAIL': df['Email'],
        'TELEFONE': df.apply(get_phone, axis=1),
        'CPF / CNPJ CLIENTE*': df.apply(lambda row: format_document(row['CPF'], row['CNPJ']), axis=1),
        'INSCR. ESTADUAL': df['Inscrição Estadual'],
        'CEP*': df['Cep'].apply(clean_cep),
        'LOGRADOURO*': df['Endereço'],
        'NÚMERO*': df['Número'],
        'COMPLEMENTO': df['Complemento'],
        'BAIRRO*': df['Bairro'],
        'CIDADE*': df['Cidade'],
        'ESTADO*': df['Estado'],
        'PONTO DE REFERÊNCIA': '',
        'VOLUMES*': '1',
        'A ENCOMENDA POSSUI NF?*': 'Sim',
        'CHAVE NF': df['Nome do Cliente'],
        'CÓDIGO INTERNO DA SUA EMPRESA (OPCIONAL)': df['Pedido'],
        'SERVIÇO DE ENVIO*': df['Frete tipo'],
        'QR CODE (Não utilizar)': '',
        'VALOR DECLARADO (OPCIONAL)': df['Subtotal produtos'],
        'OBSERVAÇÃO': df['Obs. cliente']
    })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha Mandae"

    roxo_escuro = "FF5F497A"
    rosa_claro = "FFFBE9E7"
    texto_roxo_escuro = "FF5F497A"
    branco = "FFFFFFFF"
    cinza_sem_borda = ['N', 'O', 'P', 'T', 'V']
    bordas = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fonte_titulo_branco = Font(color=branco, bold=True)
    fonte_roxo = Font(color=texto_roxo_escuro)
    alinhado_centro = Alignment(horizontal="center", vertical="center")
    alinhado_esquerda = Alignment(horizontal="left", vertical="center")
    alinhado_direita = Alignment(horizontal="right", vertical="center")

    ws.merge_cells('A1:F1')
    ws.merge_cells('G1:N1')
    ws.merge_cells('O1:R1')
    ws.merge_cells('S1:V1')
    for col in range(1, 23):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color=roxo_escuro, end_color=roxo_escuro, fill_type='solid')
        cell.font = fonte_titulo_branco
        cell.alignment = alinhado_centro
        cell.border = bordas
    ws['A1'] = "DESTINATÁRIO"
    ws['G1'] = "ENDEREÇO"
    ws['O1'] = "ENCOMENDA"
    ws['S1'] = "PEDIDO MANDAE"

    ws.append(list(saida_df.columns))
    for cell in ws[2]:
        cell.fill = PatternFill(start_color=rosa_claro, end_color=rosa_claro, fill_type='solid')
        cell.font = fonte_roxo
        cell.alignment = alinhado_centro
    ws.row_dimensions[2].height = 30

    larguras_personalizadas = {
        'A': 37.14, 'B': 30.71, 'C': 13.00, 'D': 15.71,
        'G': 12.85, 'Q': 51.42, 'V': 15.71
    }
    for col_idx in range(1, 23):
        col_letter = get_column_letter(col_idx)
        largura = larguras_personalizadas.get(col_letter, 20)
        ws.column_dimensions[col_letter].width = largura

    for i, row in saida_df.iterrows():
        ws.append(row.tolist())

    for row in ws.iter_rows(min_row=3, max_row=2+len(saida_df), min_col=1, max_col=22):
        for cell in row:
            col_letter = get_column_letter(cell.column)
            if col_letter in cinza_sem_borda:
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
            else:
                cell.fill = PatternFill(start_color=rosa_claro, end_color=rosa_claro, fill_type='solid')
                cell.border = bordas
            cell.alignment = alinhado_direita if col_letter == 'O' else alinhado_esquerda

    wb.save(output)
    output.seek(0)

    hoje = datetime.today()
    dia_util = hoje + timedelta(days=1)
    if hoje.weekday() == 4:
        dia_util += timedelta(days=2)
    nome_arquivo = f"{len(saida_df)}Pedidos - {dia_util.strftime('%d.%m')} - L2.xlsx"

    st.success("Planilha gerada com sucesso!")
    st.download_button(label="📥 Baixar Planilha Mandaê", data=output, file_name=nome_arquivo, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



import streamlit as st
import pandas as pd
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta

# 🔁 NOVA SEÇÃO
st.title("🔄 Atualizar CHAVE NF com XMLs de Nota Fiscal")
st.write("Atualize a coluna 'CHAVE NF' da planilha Mandae automaticamente com base nos arquivos XML de NF-e.")

planilha_file = st.file_uploader("1. Selecione a planilha Mandae (.xlsx):", type=["xlsx"], key="xlsx_upload")
zip_file = st.file_uploader("2. Selecione o arquivo .ZIP com os XMLs de NF-e:", type=["zip"], key="zip_upload")

if planilha_file and zip_file:
    try:
        # Abrir planilha
        wb = openpyxl.load_workbook(planilha_file)
        ws = wb.active

        # Obter índice da coluna do CPF e da CHAVE NF
        header = [cell.value for cell in ws[2]]
        idx_cpf = header.index("CPF / CNPJ CLIENTE*") + 1
        idx_chave = header.index("CHAVE NF") + 1

        # Criar dicionário de CPF -> linha
        planilha_cpfs = {}
        for row in ws.iter_rows(min_row=3, min_col=idx_cpf, max_col=idx_cpf):
            cpf_val = re.sub(r'\D', '', str(row[0].value))
            cpf_val = cpf_val.zfill(14 if len(cpf_val) > 11 else 11)
            planilha_cpfs[cpf_val] = row[0].row

        # Abrir ZIP de XMLs
        cpf_para_chave = {}
        with zipfile.ZipFile(zip_file) as z:
            for name in z.namelist():
                if name.endswith(".xml"):
                    with z.open(name) as f:
                        try:
                            tree = ET.parse(f)
                            root = tree.getroot()
                            ns = { 'ns': root.tag.split('}')[0].strip('{') }

                            doc = root.findtext('.//ns:CPF', namespaces=ns)
if not doc:
    doc = root.findtext('.//ns:CNPJ', namespaces=ns)

chave = root.findtext('.//ns:chNFe', namespaces=ns)

if doc and chave:
    doc = re.sub(r'\D', '', doc)
    doc = doc.zfill(14 if len(doc) > 11 else 11)
    cpf_para_chave[doc] = chave
                        except:
                            continue

        # Atualizar planilha
        atualizados = 0
        for cpf, row_idx in planilha_cpfs.items():
            if cpf in cpf_para_chave:
                ws.cell(row=row_idx, column=idx_chave, value=cpf_para_chave[cpf])
                atualizados += 1

        # Salvar arquivo de saída
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        hoje = datetime.today()
        dia_util = hoje + timedelta(days=1)
        if hoje.weekday() == 4:
            dia_util += timedelta(days=2)
        nome_final = f"{len(planilha_cpfs)}Pedidos - {dia_util.strftime('%d.%m')} - L2.xlsx"

        st.success(f"Chaves atualizadas com sucesso: {atualizados} de {len(planilha_cpfs)} pedidos.")
        st.download_button("📅 Baixar Planilha Atualizada", data=output, file_name=nome_final, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"Erro no processamento: {e}")
